import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Install with: pip3 install --user openpyxl")
    sys.exit(1)

wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = 'Stakeholder Analysis'

headers = ['Stakeholder', 'Role', 'Primary Goals', 'Key Queries', 'Data Needs']
ws1.append(headers)

stakeholders = [
    ['M&A Deal Team', 'Decision Makers', 'Evaluate acquisition targets\nAssess deal risk\nMake go/no-go decisions', 'Q1, Q2, Q11, Q27, Q30', 'Multi-factor risk scores\nPortfolio summaries\nFinancial impact'],
    ['Due Diligence Analysts', 'Analysts', 'Conduct property assessments\nIdentify risk factors\nGenerate reports', 'Q1, Q3, Q7, Q10, Q22', 'Property-level risk\nHistorical data\nVulnerability scores'],
    ['Portfolio Managers', 'Risk Managers', 'Assess portfolio exposure\nIdentify concentrations\nDiversification analysis', 'Q2, Q9, Q24, Q29', 'Portfolio aggregations\nGeographic clusters\nRisk distribution'],
    ['Financial Analysts', 'Financial Modeling', 'Price negotiations\nDeal structuring\nROI calculations', 'Q11, Q20, Q28', 'Financial impact\nCost-benefit analysis\nTime horizon projections'],
    ['Coastal Property Specialists', 'Domain Experts', 'Assess sea level rise\nCoastal viability\nLong-term projections', 'Q4, Q13', 'Sea level rise scenarios\nMulti-horizon projections'],
    ['Riverine Risk Analysts', 'Domain Experts', 'Assess streamflow risk\nFlood frequency\nHistorical patterns', 'Q5, Q14', 'Streamflow data\nHistorical patterns\nGauge coverage'],
    ['Model Validation Team', 'Quality Assurance', 'Validate predictions\nCompare models\nEnsure accuracy', 'Q6, Q15, Q17, Q23', 'Model performance\nData quality\nAccuracy metrics'],
    ['Insurance & Compliance', 'Regulatory', 'Determine insurance costs\nFEMA zone classification\nRegulatory requirements', 'Q7, Q12', 'FEMA zones\nInsurance requirements\nRegulatory compliance'],
    ['Post-Acquisition Team', 'Operations', 'Plan mitigation\nCost-benefit analysis\nRisk management', 'Q28', 'Mitigation strategies\nCost analysis\nROI projections'],
    ['Executive Leadership', 'Strategic Decision', 'Strategic portfolio decisions\nRisk tolerance\nInvestment strategy', 'Q16, Q30', 'Executive summaries\nComprehensive reports\nStrategic insights']
]

for row in stakeholders:
    ws1.append(row)

header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border_style
        if cell.row % 2 == 0:
            cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 35
ws1.row_dimensions[1].height = 40
for i in range(2, ws1.max_row + 1):
    ws1.row_dimensions[i].height = 60

ws2 = wb.create_sheet('Query-Stakeholder Map')
ws2.append(['Query', 'Primary Stakeholder', 'Secondary Stakeholders', 'Business Goal', 'Key Metrics'])

query_map = [
    ['Q1', 'Due Diligence Analysts', 'M&A Deal Team, Financial Analysts', 'Property acquisition viability', 'Multi-factor risk score, Financial impact'],
    ['Q2', 'Portfolio Managers', 'M&A Deal Team, Executive Leadership', 'Deal risk evaluation', 'Portfolio risk score, Geographic concentration'],
    ['Q3', 'Due Diligence Analysts', 'M&A Deal Team, Financial Analysts', 'Historical risk assessment', 'Flood frequency, Recurrence intervals'],
    ['Q4', 'Coastal Property Specialists', 'M&A Deal Team, Financial Analysts', 'Long-term coastal viability', 'Sea level rise projections, Scenario comparison'],
    ['Q5', 'Riverine Risk Analysts', 'Due Diligence Analysts', 'Riverine flood risk', 'Streamflow frequency, Historical patterns'],
    ['Q6', 'Model Validation Team', 'All Stakeholders', 'Prediction reliability', 'Model accuracy, Performance metrics'],
    ['Q7', 'Insurance & Compliance', 'Due Diligence Analysts', 'Regulatory compliance', 'FEMA zone classification, Insurance costs'],
    ['Q8', 'Portfolio Managers', 'M&A Deal Team', 'Risk trend analysis', 'Temporal trends, Risk direction'],
    ['Q9', 'Portfolio Managers', 'M&A Deal Team', 'Geographic clustering', 'Risk clusters, Concentration metrics'],
    ['Q10', 'Due Diligence Analysts', 'M&A Deal Team', 'Vulnerability prioritization', 'Vulnerability scores, Risk ranking'],
    ['Q11', 'Financial Analysts', 'M&A Deal Team', 'Price negotiation', 'Financial impact, Deal terms'],
    ['Q12', 'Insurance & Compliance', 'Due Diligence Analysts', 'Insurance cost determination', 'FEMA zone levels, Regulatory requirements'],
    ['Q13', 'Coastal Property Specialists', 'Financial Analysts', 'Scenario comparison', 'Best/worst case outcomes'],
    ['Q14', 'Riverine Risk Analysts', 'Due Diligence Analysts', 'Future frequency prediction', 'Historical patterns, Frequency trends'],
    ['Q15', 'Model Validation Team', 'All Stakeholders', 'NASA model accuracy', 'Prediction reliability, Model performance'],
    ['Q16', 'Executive Leadership', 'Portfolio Managers', 'Stakeholder presentation', 'Portfolio summaries, Risk metrics'],
    ['Q17', 'Model Validation Team', 'All Stakeholders', 'Data quality assurance', 'Quality metrics, Confidence levels'],
    ['Q18', 'Due Diligence Analysts', 'Portfolio Managers', 'Performance optimization', 'Query performance, Large portfolio handling'],
    ['Q19', 'M&A Deal Team', 'All Stakeholders', 'Unified risk rating', 'Combined scores, Simplified metrics'],
    ['Q20', 'Financial Analysts', 'M&A Deal Team', 'Time horizon assessment', 'Multi-horizon projections, Property value'],
    ['Q21', 'Due Diligence Analysts', 'Post-Acquisition Team', 'Mitigation needs', 'Elevation protection, Mitigation requirements'],
    ['Q22', 'Due Diligence Analysts', 'Financial Analysts', 'Future loss estimation', 'Historical damage, Loss projections'],
    ['Q23', 'Model Validation Team', 'All Stakeholders', 'Model selection', 'Accuracy comparison, Best model'],
    ['Q24', 'Portfolio Managers', 'Executive Leadership', 'Diversification analysis', 'Geographic distribution, Concentration'],
    ['Q25', 'Portfolio Managers', 'M&A Deal Team', 'Asset class risk', 'Property type risk, Vulnerability by type'],
    ['Q26', 'Due Diligence Analysts', 'Portfolio Managers', 'Compound risk scenarios', 'Cascading effects, Compound risk'],
    ['Q27', 'M&A Deal Team', 'Due Diligence Analysts', 'Deal-breaker identification', 'Risk thresholds, Property filtering'],
    ['Q28', 'Post-Acquisition Team', 'Financial Analysts', 'Mitigation cost-benefit', 'ROI analysis, Cost-effectiveness'],
    ['Q29', 'Portfolio Managers', 'Executive Leadership', 'Diversification assessment', 'Risk concentration, Geographic spread'],
    ['Q30', 'Executive Leadership', 'M&A Deal Team', 'Final decision support', 'Comprehensive report, Go/no-go decision']
]

for row in query_map:
    ws2.append(row)

for cell in ws2[1]:
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border_style
        if cell.row % 2 == 0:
            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 30
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['E'].width = 30
ws2.row_dimensions[1].height = 40
for i in range(2, ws2.max_row + 1):
    ws2.row_dimensions[i].height = 50

ws3 = wb.create_sheet('Data Requirements')
ws3.append(['Data Source', 'Stakeholder Needs', 'Queries Using', 'Critical Fields', 'Update Frequency'])

data_req = [
    ['FEMA Flood Zones', 'Insurance costs, Regulatory compliance, Zone classification', 'Q1, Q7, Q12', 'zone_code, base_flood_elevation, zone_geom', 'Annual'],
    ['NOAA Sea Level Rise', 'Coastal viability, Long-term projections, Scenario comparison', 'Q1, Q4, Q13', 'projection_year, scenario, sea_level_rise_feet', 'Quarterly'],
    ['USGS Streamflow', 'Riverine risk, Flood frequency, Historical patterns', 'Q1, Q5, Q14', 'gauge_id, discharge_cfs, flood_category', 'Real-time'],
    ['NASA Flood Models', 'Prediction accuracy, Model validation, Flood probability', 'Q1, Q6, Q15, Q23', 'flood_probability, inundation_depth_feet, model_id', 'Daily'],
    ['Historical Flood Events', 'Past damage, Frequency patterns, Recurrence intervals', 'Q3, Q22', 'event_date, severity, damage_estimate_dollars', 'As available'],
    ['Property Data', 'Property characteristics, Location, Value', 'All queries', 'property_id, property_geom, total_value, elevation_feet', 'Per acquisition'],
    ['Portfolio Data', 'Portfolio structure, Geographic distribution', 'Q2, Q9, Q16, Q24, Q29', 'portfolio_id, portfolio_name, acquisition_date', 'Per acquisition']
]

for row in data_req:
    ws3.append(row)

for cell in ws3[1]:
    cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    cell.font = Font(bold=True, color='000000', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = border_style
        if cell.row % 2 == 0:
            cell.fill = PatternFill(start_color='FFF4CC', end_color='FFF4CC', fill_type='solid')

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 35
ws3.column_dimensions['E'].width = 15
ws3.row_dimensions[1].height = 40
for i in range(2, ws3.max_row + 1):
    ws3.row_dimensions[i].height = 60

ws4 = wb.create_sheet('Query Priority Matrix')
ws4.append(['Query', 'Business Impact', 'Stakeholder Count', 'Data Complexity', 'Priority Score', 'Development Phase'])

priority_data = [
    ['Q1', 'Critical', 4, 'Very High', 95, 'Phase 1'],
    ['Q2', 'Critical', 3, 'Very High', 90, 'Phase 1'],
    ['Q11', 'Critical', 2, 'High', 85, 'Phase 1'],
    ['Q30', 'Critical', 2, 'Very High', 85, 'Phase 1'],
    ['Q27', 'High', 2, 'Medium', 80, 'Phase 1'],
    ['Q3', 'High', 3, 'High', 75, 'Phase 2'],
    ['Q4', 'High', 2, 'High', 75, 'Phase 2'],
    ['Q7', 'High', 2, 'Medium', 70, 'Phase 2'],
    ['Q10', 'High', 2, 'Medium', 70, 'Phase 2'],
    ['Q12', 'High', 2, 'Low', 70, 'Phase 2'],
    ['Q16', 'High', 2, 'Medium', 70, 'Phase 2'],
    ['Q5', 'Medium', 2, 'High', 65, 'Phase 3'],
    ['Q6', 'Medium', 1, 'Medium', 65, 'Phase 3'],
    ['Q8', 'Medium', 2, 'Medium', 65, 'Phase 3'],
    ['Q9', 'Medium', 2, 'High', 65, 'Phase 3'],
    ['Q13', 'Medium', 2, 'High', 65, 'Phase 3'],
    ['Q14', 'Medium', 2, 'High', 65, 'Phase 3'],
    ['Q15', 'Medium', 1, 'Medium', 65, 'Phase 3'],
    ['Q17', 'Medium', 1, 'Low', 65, 'Phase 3'],
    ['Q19', 'Medium', 1, 'Medium', 65, 'Phase 3'],
    ['Q20', 'Medium', 2, 'High', 65, 'Phase 3'],
    ['Q21', 'Medium', 2, 'Medium', 65, 'Phase 3'],
    ['Q22', 'Medium', 2, 'Medium', 65, 'Phase 3'],
    ['Q23', 'Medium', 1, 'Medium', 65, 'Phase 3'],
    ['Q24', 'Medium', 2, 'High', 65, 'Phase 3'],
    ['Q25', 'Medium', 2, 'Medium', 65, 'Phase 3'],
    ['Q26', 'Medium', 2, 'Very High', 65, 'Phase 3'],
    ['Q28', 'Medium', 2, 'Medium', 60, 'Phase 4'],
    ['Q18', 'Low', 2, 'Medium', 55, 'Phase 4'],
    ['Q29', 'Low', 2, 'High', 55, 'Phase 4']
]

for row in priority_data:
    ws4.append(row)

for cell in ws4[1]:
    cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_style
        if cell.column == 2:
            if cell.value == 'Critical':
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            elif cell.value == 'High':
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        elif cell.column == 6:
            if cell.value == 'Phase 1':
                cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            elif cell.value == 'Phase 2':
                cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            elif cell.value == 'Phase 3':
                cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
        elif cell.row % 2 == 0:
            if cell.column not in [2, 6]:
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

ws4.column_dimensions['A'].width = 8
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 15
ws4.column_dimensions['F'].width = 15
ws4.row_dimensions[1].height = 40
for i in range(2, ws4.max_row + 1):
    ws4.row_dimensions[i].height = 30

output_path = os.path.join(os.path.dirname(__file__), '..', 'docs', 'M&A_Stakeholder_Analysis.xlsx')
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f'Excel file created successfully: {output_path}')
